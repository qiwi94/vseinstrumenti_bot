import openpyxl
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackContext

TOKEN = '7475740177:AAEjeLBqPK4j8h6KMVw8nKMEAoTSMqJPsNs' 

# Загрузка книги Excel
try:
    book = openpyxl.load_workbook('123.xlsx')

    # Проверяем, что листы 'День' и 'Ночь' существуют
    available_sheets = book.sheetnames
    if 'День' not in available_sheets or 'Ночь' not in available_sheets:
        raise KeyError('Ошибка: Один или оба листа "День" и "Ночь" отсутствуют в книге.')

    sheet_day = book['День']
    sheet_night = book['Ночь']
except FileNotFoundError:
    print("Ошибка: файл '123.xlsx' не найден.")
    exit()
except KeyError as e:
    print(e)
    exit()

# Создание словарей для листов "День" и "Ночь"
list_day = {sheet_day.cell(row=i, column=1).value.lower(): i for i in range(3, sheet_day.max_row + 1)}
list_night = {sheet_night.cell(row=i, column=1).value.lower(): i for i in range(3, sheet_night.max_row + 1)}


def is_name(text, lst):
    return text in lst


def is_not_null(x):
    return x is not None


def is_int(x):
    return isinstance(x, (int, float))


def count_many_day(text, num_name_day):
    result = []
    if is_name(text, list_day):
        for i in range(2, sheet_day.max_column + 1):
            if is_not_null(sheet_day.cell(row=num_name_day, column=i).value) and is_int(
                    sheet_day.cell(row=num_name_day, column=i).value):
                result.append(sheet_day.cell(row=num_name_day, column=i).value * sheet_day.cell(row=3, column=i).value)
        return round(sum(result), 2)
    return 0


def count_many_night(text, num_name_night):
    result = []
    if is_name(text, list_night):
        for i in range(2, sheet_night.max_column + 1):
            if is_not_null(sheet_night.cell(row=num_name_night, column=i).value) and is_int(
                    sheet_night.cell(row=num_name_night, column=i).value):
                result.append(
                    sheet_night.cell(row=num_name_night, column=i).value * sheet_night.cell(row=3, column=i).value)
        return round(sum(result), 2)
    return 0


def sum_may(f1, f2):
    return round(f1 + f2, 2)


async def start(update: Update, context: CallbackContext):
    await update.message.reply_text(
        "Привет! Введите команду /salary и ваше ФИО, чтобы узнать ваш доход.\nПример: /salary Иванов Иван Иванович")


async def salary(update: Update, context: CallbackContext):
    if len(context.args) < 2:
        await update.message.reply_text("Пожалуйста, введите ваше ФИО.")
        return

    name = ' '.join(context.args).lower()

    try:
        num_name_day = list_day[name]
        num_name_night = list_night[name]
    except KeyError:
        await update.message.reply_text(f"'{name}' - введенное ФИО отсутствует в списке! Попробуйте ещё раз!")
        return

    day = count_many_day(name, num_name_day)
    night = count_many_night(name, num_name_night)
    full_many = sum_may(day, night)

    await update.message.reply_text(
        f"Доход за дневные часы = {day} ₽\n"
        f"Доход за ночные часы = {night} ₽\n"
        f"Общий доход = {full_many} ₽"
    )


def main():
    application = Application.builder().token(TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("salary", salary))

    print("Бот запущен. Нажмите Ctrl+C для завершения.")
    application.run_polling()


if __name__ == '__main__':
    main()
